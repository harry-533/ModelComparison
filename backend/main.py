from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from ultralytics import YOLO
from google.cloud import vision
from PIL import Image, ImageDraw
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from typing import List
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import io
import requests
import shutil
import os, base64
import cv2
import numpy as np
import base64
import uuid

app = FastAPI()

# Allow frontend (running on file:// or localhost)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # For development only
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")
app.mount("/downloads", StaticFiles(directory="frontend/downloads"), name="downloads")

if "GOOGLE_APPLICATION_CREDENTIALS_JSON" in os.environ:
    creds_data = base64.b64decode(os.environ["GOOGLE_APPLICATION_CREDENTIALS_JSON"])
    with open("google_creds.json", "wb") as f:
        f.write(creds_data)
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_creds.json"


if not os.path.exists("yolo.pt"):
    print("Downloading YOLO model from Google Drive...")
    os.system("gdown --id 1tFAN0ies3wIsLC4q--PGMR8MoWAWRLNU -O yolo.pt")

model = YOLO('yolo.pt')

@app.post("/upload/")
async def upload_image(file: UploadFile = File(...)):
    # Save the uploaded file temporarily
    file_path = f"temp_{file.filename}"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    google_labels, google_objects = detect_labels_from_image_path(file_path)

    results = model(file_path)[0]
    img = cv2.imread(file_path)
    yolo_labels = []

    for box in results.boxes:
        cls_id = int(box.cls[0])
        conf = float(box.conf[0])
        label = model.names[cls_id]
        x1, y1, x2, y2 = map(int, box.xyxy[0])

        # Draw box and label on image
        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(img, f"{label} {conf:.2f}", (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        
        yolo_labels.append({
            "label": label,
            "confidence": round(conf, 4),
        })

    _, buffer_yolo = cv2.imencode('.jpg', img)
    img_yolo_base64 = base64.b64encode(buffer_yolo).decode('utf-8')
    
    # ===== Google Drawing Section =====
    img_google = Image.open(file_path)
    draw = ImageDraw.Draw(img_google)

    for obj in google_objects:
        box = [(v.x * img_google.width, v.y * img_google.height) for v in obj.bounding_poly.normalized_vertices]
        draw.line(box + [box[0]], width=3, fill='red')
        draw.text(box[0], obj.name, fill='white')

    img_google_bytes = io.BytesIO()
    img_google.save(img_google_bytes, format='JPEG')
    img_google_base64 = base64.b64encode(img_google_bytes.getvalue()).decode('utf-8')

    # Cleanup temp file
    os.remove(file_path)

    return JSONResponse(content={
        "yolo_labels": yolo_labels,
        "google_labels": google_labels,
        "yolo_image_base64": img_yolo_base64,
        "google_image_base64": img_google_base64
    })

@app.post("/upload-folder/")
async def upload_folder(images: list[UploadFile] = File(...)):
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "YOLO Results"

    row = 1

    for file in images:
        clean_name = os.path.basename(file.filename)
        filename = f"temp_{uuid.uuid4().hex[:8]}_{clean_name}"
        with open(filename, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        results = model(filename)[0]
        labels = []

        for box in results.boxes:
            cls_id = int(box.cls[0])
            conf = float(box.conf[0])
            label = model.names[cls_id]
            labels.append((label, round(conf, 4)))

        # Save image with boxes
        img = cv2.imread(filename)
        for box in results.boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            label = model.names[int(box.cls[0])]
            conf = float(box.conf[0])
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(img, f"{label} {conf:.2f}", (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

        output_img_path = f"annotated_{uuid.uuid4().hex[:8]}.jpg"
        cv2.imwrite(output_img_path, img)

        # Insert Image
        if os.path.exists(output_img_path):
            ws.row_dimensions[row].height = 100
            img_excel = ExcelImage(output_img_path)
            img_excel.width = 200
            img_excel.height = 100
            ws.add_image(img_excel, f"A{row}")
        else:
            print(f"Warning: Annotated image {output_img_path} not found. Skipping Excel embedding.")

        # Write labels
        ws.cell(row=row, column=2, value="Filename")
        ws.cell(row=row, column=3, value=file.filename)
        row += 1
        ws.cell(row=row, column=2, value="Label")
        ws.cell(row=row, column=3, value="Confidence")
        row += 1

        for label, conf in labels:
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=conf)
            row += 1

        # Spacer row
        row += 1

    # Save Excel
    excel_path = f"frontend/downloads/results_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(excel_path)

    for file in os.listdir():
        if file.startswith("temp_") or file.startswith("annotated_"):
            os.remove(file)

    return {"download_url": f"{excel_path}"}

def detect_labels_from_image_path(image_path: str):
    client = vision.ImageAnnotatorClient()

    with io.open(image_path, 'rb') as image_file:
        content = image_file.read()

    image = vision.Image(content=content)

    # Label Detection (no bounding boxes)
    label_response = client.label_detection(image=image)
    labels = [
        {"label": label.description, "confidence": round(label.score, 4)}
        for label in label_response.label_annotations
    ]

    # Object Detection (bounding boxes)
    object_response = client.object_localization(image=image)
    objects = object_response.localized_object_annotations

    return labels, objects

def draw_google_boxes(image_path, objects):
    image = Image.open(image_path)
    draw = ImageDraw.Draw(image)

    for obj in objects:
        # Normalized coordinates to pixels
        box = [(v.x * image.width, v.y * image.height) for v in obj.bounding_poly.normalized_vertices]
        draw.line(box + [box[0]], width=3, fill='red')
        draw.text(box[0], obj.name, fill='white')

    # Save as labeled image
    output_name = f"labeled_{uuid.uuid4().hex[:8]}.jpg"
    output_path = os.path.join("labeled", output_name)
    os.makedirs("labeled", exist_ok=True)
    image.save(output_path)

    return output_name